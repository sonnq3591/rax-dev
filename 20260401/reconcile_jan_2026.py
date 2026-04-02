from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path("/workspaces/rax-dev/20260401")
OUT_DIR = BASE_DIR / "jan_2026_reconciliation"
FILES = {
    "pbi": BASE_DIR / "pbi.xlsx",
    "synapse": BASE_DIR / "synapse.xlsx",
}

COMPANY = "Company code.Company code Level 01.Company (Key)"
REPORTING_DATE = "Reporting date"
LEVEL3 = "G/L Account.Level 03.Key"
LEVEL4 = "G/L Account.Level 04.Key"
ACTUAL = "MTD&0T_FPIE&Actual"
BUDGET = "MTD&0T_FPIE&Budget"

TARGET_MONTH = "2026-01"
SHARED_COMPANIES = {"EA01", "EA11", "EA12", "EA18", "EA22", "EA37"}


def normalize_date(value: object) -> str:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.isoformat()
    return "" if value is None else str(value).strip()


def numeric(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value.upper() == "NULL":
            return Decimal("0")
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal("0")


def normalize_key(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text == "" or text.upper() == "NULL":
        return ""
    return text


def fmt_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f")


def load_filtered_aggregates() -> tuple[dict[str, dict[str, dict[str, Decimal | int]]], dict[str, dict[tuple[str, str], dict[str, Decimal | int]]], dict[str, dict[tuple[str, str], dict[str, Decimal | int]]]]:
    company_totals = {label: defaultdict(lambda: {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")}) for label in FILES}
    level3_totals = {label: defaultdict(lambda: {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")}) for label in FILES}
    level4_totals = {label: defaultdict(lambda: {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")}) for label in FILES}

    for label, path in FILES.items():
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        idx = {header: i for i, header in enumerate(headers)}

        for row in rows:
            company = row[idx[COMPANY]]
            if company not in SHARED_COMPANIES:
                continue

            reporting_date = normalize_date(row[idx[REPORTING_DATE]])
            if not reporting_date.startswith(TARGET_MONTH):
                continue

            actual = numeric(row[idx[ACTUAL]])
            budget = numeric(row[idx[BUDGET]])

            company_record = company_totals[label][company]
            company_record["rows"] += 1
            company_record["actual"] += actual
            company_record["budget"] += budget

            level3_key = (company, normalize_key(row[idx[LEVEL3]]))
            level3_record = level3_totals[label][level3_key]
            level3_record["rows"] += 1
            level3_record["actual"] += actual
            level3_record["budget"] += budget

            level4_key = (company, normalize_key(row[idx[LEVEL4]]))
            level4_record = level4_totals[label][level4_key]
            level4_record["rows"] += 1
            level4_record["actual"] += actual
            level4_record["budget"] += budget

    return company_totals, level3_totals, level4_totals


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def company_summary(company_totals: dict[str, dict[str, dict[str, Decimal | int]]]) -> list[dict[str, object]]:
    rows = []
    for company in sorted(SHARED_COMPANIES):
        pbi = company_totals["pbi"].get(company, {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        synapse = company_totals["synapse"].get(company, {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        rows.append(
            {
                "company_code": company,
                "pbi_rows": pbi["rows"],
                "synapse_rows": synapse["rows"],
                "row_diff": pbi["rows"] - synapse["rows"],
                "pbi_actual": fmt_decimal(pbi["actual"]),
                "synapse_actual": fmt_decimal(synapse["actual"]),
                "actual_diff": fmt_decimal(pbi["actual"] - synapse["actual"]),
                "pbi_budget": fmt_decimal(pbi["budget"]),
                "synapse_budget": fmt_decimal(synapse["budget"]),
                "budget_diff": fmt_decimal(pbi["budget"] - synapse["budget"]),
            }
        )
    return rows


def diff_rows(agg: dict[str, dict[tuple[str, str], dict[str, Decimal | int]]], level_name: str) -> list[dict[str, object]]:
    results = []
    all_keys = sorted(set(agg["pbi"]) | set(agg["synapse"]))
    for company, level_key in all_keys:
        pbi = agg["pbi"].get((company, level_key), {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        synapse = agg["synapse"].get((company, level_key), {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        row_diff = pbi["rows"] - synapse["rows"]
        actual_diff = pbi["actual"] - synapse["actual"]
        budget_diff = pbi["budget"] - synapse["budget"]
        if row_diff == 0 and actual_diff == 0 and budget_diff == 0:
            continue
        results.append(
            {
                "company_code": company,
                level_name: level_key,
                "pbi_rows": pbi["rows"],
                "synapse_rows": synapse["rows"],
                "row_diff": row_diff,
                "pbi_actual": fmt_decimal(pbi["actual"]),
                "synapse_actual": fmt_decimal(synapse["actual"]),
                "actual_diff": fmt_decimal(actual_diff),
                "pbi_budget": fmt_decimal(pbi["budget"]),
                "synapse_budget": fmt_decimal(synapse["budget"]),
                "budget_diff": fmt_decimal(budget_diff),
            }
        )
    results.sort(
        key=lambda row: (
            abs(Decimal(str(row["actual_diff"]))),
            abs(Decimal(str(row["budget_diff"]))),
            abs(int(row["row_diff"])),
        ),
        reverse=True,
    )
    return results


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    company_totals, level3_totals, level4_totals = load_filtered_aggregates()

    company_rows = company_summary(company_totals)
    level3_rows = diff_rows(level3_totals, "level3_key")
    level4_rows = diff_rows(level4_totals, "level4_key")

    write_csv(
        OUT_DIR / "company_summary.csv",
        company_rows,
        [
            "company_code",
            "pbi_rows",
            "synapse_rows",
            "row_diff",
            "pbi_actual",
            "synapse_actual",
            "actual_diff",
            "pbi_budget",
            "synapse_budget",
            "budget_diff",
        ],
    )
    write_csv(
        OUT_DIR / "level3_diffs.csv",
        level3_rows,
        [
            "company_code",
            "level3_key",
            "pbi_rows",
            "synapse_rows",
            "row_diff",
            "pbi_actual",
            "synapse_actual",
            "actual_diff",
            "pbi_budget",
            "synapse_budget",
            "budget_diff",
        ],
    )
    write_csv(
        OUT_DIR / "level4_diffs.csv",
        level4_rows,
        [
            "company_code",
            "level4_key",
            "pbi_rows",
            "synapse_rows",
            "row_diff",
            "pbi_actual",
            "synapse_actual",
            "actual_diff",
            "pbi_budget",
            "synapse_budget",
            "budget_diff",
        ],
    )

    summary = {
        "target_month": TARGET_MONTH,
        "shared_companies": sorted(SHARED_COMPANIES),
        "company_summary_preview": company_rows,
        "level3_diff_count": len(level3_rows),
        "level4_diff_count": len(level4_rows),
        "top_level3_diffs": level3_rows[:20],
        "top_level4_diffs": level4_rows[:20],
    }
    (OUT_DIR / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
