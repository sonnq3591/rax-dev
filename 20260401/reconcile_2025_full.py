from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path("/workspaces/rax-dev/20260401")
OUT_DIR = BASE_DIR / "reconciliation_2025_full"
FILES = {
    "pbi": BASE_DIR / "pbi.xlsx",
    "synapse": BASE_DIR / "synapse.xlsx",
}

COMPANY = "Company code.Company code Level 01.Company (Key)"
REPORTING_DATE = "Reporting date"
LEVEL3 = "G/L Account.Level 03.Key"
LEVEL4 = "G/L Account.Level 04.Key"
ACTUAL = "MTD&0T_FPIE&Actual"
BUDGET = "MTD&0T_FPIE&Budget"

TARGET_YEAR = "2025"
SHARED_COMPANIES = {"EA01", "EA11", "EA12", "EA18", "EA22", "EA37"}


def normalize_date(value: object) -> str:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.isoformat()
    return "" if value is None else str(value).strip()


def month_key(value: object) -> str:
    text = normalize_date(value)
    return text[:7] if len(text) >= 7 else text


def numeric(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value.upper() == "NULL":
            return Decimal("0")
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal("0")


def normalize_key(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text == "" or text.upper() == "NULL":
        return ""
    return text


def fmt_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f")


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_aggregates():
    company_month = {label: defaultdict(lambda: {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")}) for label in FILES}
    level3 = {label: defaultdict(lambda: {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")}) for label in FILES}
    level4 = {label: defaultdict(lambda: {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")}) for label in FILES}

    for label, path in FILES.items():
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        idx = {header: i for i, header in enumerate(headers)}

        for row in rows:
            company = row[idx[COMPANY]]
            if company not in SHARED_COMPANIES:
                continue

            month = month_key(row[idx[REPORTING_DATE]])
            if not month.startswith(TARGET_YEAR):
                continue

            actual = numeric(row[idx[ACTUAL]])
            budget = numeric(row[idx[BUDGET]])

            company_month_key = (company, month)
            company_month_rec = company_month[label][company_month_key]
            company_month_rec["rows"] += 1
            company_month_rec["actual"] += actual
            company_month_rec["budget"] += budget

            level3_key = (company, month, normalize_key(row[idx[LEVEL3]]))
            level3_rec = level3[label][level3_key]
            level3_rec["rows"] += 1
            level3_rec["actual"] += actual
            level3_rec["budget"] += budget

            level4_key = (company, month, normalize_key(row[idx[LEVEL4]]))
            level4_rec = level4[label][level4_key]
            level4_rec["rows"] += 1
            level4_rec["actual"] += actual
            level4_rec["budget"] += budget

    return company_month, level3, level4


def diff_company_month(company_month):
    rows = []
    for company, month in sorted(set(company_month["pbi"]) | set(company_month["synapse"])):
        pbi = company_month["pbi"].get((company, month), {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        synapse = company_month["synapse"].get((company, month), {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        row_diff = pbi["rows"] - synapse["rows"]
        actual_diff = pbi["actual"] - synapse["actual"]
        budget_diff = pbi["budget"] - synapse["budget"]
        rows.append(
            {
                "company_code": company,
                "month": month,
                "pbi_rows": pbi["rows"],
                "synapse_rows": synapse["rows"],
                "row_diff": row_diff,
                "pbi_actual": fmt_decimal(pbi["actual"]),
                "synapse_actual": fmt_decimal(synapse["actual"]),
                "actual_diff": fmt_decimal(actual_diff),
                "pbi_budget": fmt_decimal(pbi["budget"]),
                "synapse_budget": fmt_decimal(synapse["budget"]),
                "budget_diff": fmt_decimal(budget_diff),
                "matches": row_diff == 0 and actual_diff == 0 and budget_diff == 0,
            }
        )
    return rows


def diff_deep(agg, level_name: str, failing_company_months: set[tuple[str, str]]):
    rows = []
    for company, month, key in sorted(set(agg["pbi"]) | set(agg["synapse"])):
        if (company, month) not in failing_company_months:
            continue
        pbi = agg["pbi"].get((company, month, key), {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        synapse = agg["synapse"].get((company, month, key), {"rows": 0, "actual": Decimal("0"), "budget": Decimal("0")})
        row_diff = pbi["rows"] - synapse["rows"]
        actual_diff = pbi["actual"] - synapse["actual"]
        budget_diff = pbi["budget"] - synapse["budget"]
        if row_diff == 0 and actual_diff == 0 and budget_diff == 0:
            continue
        rows.append(
            {
                "company_code": company,
                "month": month,
                level_name: key,
                "pbi_rows": pbi["rows"],
                "synapse_rows": synapse["rows"],
                "row_diff": row_diff,
                "pbi_actual": fmt_decimal(pbi["actual"]),
                "synapse_actual": fmt_decimal(synapse["actual"]),
                "actual_diff": fmt_decimal(actual_diff),
                "pbi_budget": fmt_decimal(pbi["budget"]),
                "synapse_budget": fmt_decimal(synapse["budget"]),
                "budget_diff": fmt_decimal(budget_diff),
            }
        )
    rows.sort(
        key=lambda row: (
            abs(Decimal(str(row["actual_diff"]))),
            abs(Decimal(str(row["budget_diff"]))),
            abs(int(row["row_diff"])),
        ),
        reverse=True,
    )
    return rows


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    company_month, level3, level4 = load_aggregates()

    company_month_rows = diff_company_month(company_month)
    failing = {
        (row["company_code"], row["month"])
        for row in company_month_rows
        if not row["matches"]
    }
    level3_rows = diff_deep(level3, "level3_key", failing)
    level4_rows = diff_deep(level4, "level4_key", failing)

    write_csv(
        OUT_DIR / "company_month_summary.csv",
        company_month_rows,
        [
            "company_code",
            "month",
            "pbi_rows",
            "synapse_rows",
            "row_diff",
            "pbi_actual",
            "synapse_actual",
            "actual_diff",
            "pbi_budget",
            "synapse_budget",
            "budget_diff",
            "matches",
        ],
    )
    write_csv(
        OUT_DIR / "level3_diffs_for_failed_months.csv",
        level3_rows,
        [
            "company_code",
            "month",
            "level3_key",
            "pbi_rows",
            "synapse_rows",
            "row_diff",
            "pbi_actual",
            "synapse_actual",
            "actual_diff",
            "pbi_budget",
            "synapse_budget",
            "budget_diff",
        ],
    )
    write_csv(
        OUT_DIR / "level4_diffs_for_failed_months.csv",
        level4_rows,
        [
            "company_code",
            "month",
            "level4_key",
            "pbi_rows",
            "synapse_rows",
            "row_diff",
            "pbi_actual",
            "synapse_actual",
            "actual_diff",
            "pbi_budget",
            "synapse_budget",
            "budget_diff",
        ],
    )

    summary = {
        "target_year": TARGET_YEAR,
        "shared_companies": sorted(SHARED_COMPANIES),
        "company_month_count": len(company_month_rows),
        "failed_company_month_count": len(failing),
        "failed_company_months": [
            {
                "company_code": company,
                "month": month,
            }
            for company, month in sorted(failing)
        ],
        "company_month_preview": company_month_rows[:24],
        "top_level3_diffs": level3_rows[:25],
        "top_level4_diffs": level4_rows[:25],
    }
    (OUT_DIR / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
