from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


BASE_DIR = Path("/workspaces/rax-dev/20260401")
PBI_PATH = BASE_DIR / "pbi.xlsx"
SYNAPSE_PATH = BASE_DIR / "synapse.xlsx"
OUT_DIR = BASE_DIR / "comparison_output"

COMPANY_KEY = "Company code.Company code Level 01.Company (Key)"
LEVEL3_KEY = "G/L Account.Level 03.Key"
LEVEL4_KEY = "G/L Account.Level 04.Key"
MEASURE_COLUMNS = ["MTD&0T_FPIE&Actual", "MTD&0T_FPIE&Budget"]
PREVIEW_LIMIT = 200


@dataclass
class WorkbookData:
    label: str
    path: Path
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, object]]


def trim_headers(values: Iterable[object]) -> list[str]:
    headers = ["" if value is None else str(value).strip() for value in values]
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def normalize_blank(value: object) -> object | None:
    if value is None:
        return None
    if isinstance(value, str):
        trimmed = value.strip()
        if trimmed == "" or trimmed.upper() == "NULL":
            return None
        return trimmed
    return value


def normalize_value(value: object) -> str:
    value = normalize_blank(value)
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float, Decimal)):
        try:
            return format(Decimal(str(value)).normalize(), "f")
        except InvalidOperation:
            return str(value)
    return str(value)


def numeric_value(value: object) -> Decimal:
    value = normalize_blank(value)
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal("0")


def load_workbook_data(label: str, path: Path) -> WorkbookData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    row_iter = sheet.iter_rows(values_only=True)
    headers = trim_headers(next(row_iter))
    rows: list[dict[str, object]] = []
    header_count = len(headers)

    for values in row_iter:
        values = tuple(values[:header_count])
        if not any(value is not None and str(value).strip() != "" for value in values):
            continue
        row = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(header_count)}
        rows.append(row)

    return WorkbookData(
        label=label,
        path=path,
        sheet_name=sheet.title,
        headers=headers,
        rows=rows,
    )


def schema_report(left: WorkbookData, right: WorkbookData) -> dict[str, object]:
    left_only = [header for header in left.headers if header not in right.headers]
    right_only = [header for header in right.headers if header not in left.headers]
    shared = [header for header in left.headers if header in right.headers]

    order_differences = []
    for index, header in enumerate(shared):
        right_index = right.headers.index(header)
        if right_index != index:
            order_differences.append(
                {
                    "column": header,
                    "pbi_position": index + 1,
                    "synapse_position": right_index + 1,
                }
            )

    return {
        "pbi_sheet": left.sheet_name,
        "synapse_sheet": right.sheet_name,
        "pbi_column_count": len(left.headers),
        "synapse_column_count": len(right.headers),
        "shared_column_count": len(shared),
        "columns_only_in_pbi": left_only,
        "columns_only_in_synapse": right_only,
        "column_order_differences": order_differences,
        "shared_columns": shared,
    }


def row_signature(row: dict[str, object], columns: list[str]) -> tuple[str, ...]:
    return tuple(normalize_value(row.get(column)) for column in columns)


def aggregate_by_keys(
    rows: list[dict[str, object]], key_columns: list[str], measure_columns: list[str]
) -> dict[tuple[str, ...], dict[str, object]]:
    groups: dict[tuple[str, ...], dict[str, object]] = {}
    for row in rows:
        key = tuple(normalize_value(row.get(column)) for column in key_columns)
        group = groups.setdefault(
            key,
            {
                "row_count": 0,
                "measures": defaultdict(lambda: Decimal("0")),
            },
        )
        group["row_count"] += 1
        measures = group["measures"]
        for measure in measure_columns:
            measures[measure] += numeric_value(row.get(measure))
    return groups


def format_group_record(key_columns: list[str], key: tuple[str, ...], stats: dict[str, object]) -> dict[str, object]:
    record = {column: value for column, value in zip(key_columns, key)}
    record["row_count"] = stats["row_count"]
    for measure, total in stats["measures"].items():
        record[measure] = format(total.normalize(), "f")
    return record


def write_csv(path: Path, rows: list[dict[str, object]], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def diff_groupings(
    left: WorkbookData,
    right: WorkbookData,
    key_columns: list[str],
    measure_columns: list[str],
    output_prefix: str,
) -> dict[str, object]:
    left_groups = aggregate_by_keys(left.rows, key_columns, measure_columns)
    right_groups = aggregate_by_keys(right.rows, key_columns, measure_columns)

    only_left: list[dict[str, object]] = []
    only_right: list[dict[str, object]] = []
    changed: list[dict[str, object]] = []

    all_keys = sorted(set(left_groups) | set(right_groups))
    for key in all_keys:
        left_stats = left_groups.get(key)
        right_stats = right_groups.get(key)
        if left_stats and not right_stats:
            only_left.append(format_group_record(key_columns, key, left_stats))
            continue
        if right_stats and not left_stats:
            only_right.append(format_group_record(key_columns, key, right_stats))
            continue

        assert left_stats and right_stats
        measure_diffs = {}
        for measure in measure_columns:
            left_total = left_stats["measures"][measure]
            right_total = right_stats["measures"][measure]
            if left_total != right_total:
                measure_diffs[measure] = {
                    "pbi": format(left_total.normalize(), "f"),
                    "synapse": format(right_total.normalize(), "f"),
                }

        if left_stats["row_count"] != right_stats["row_count"] or measure_diffs:
            record = {column: value for column, value in zip(key_columns, key)}
            record["pbi_row_count"] = left_stats["row_count"]
            record["synapse_row_count"] = right_stats["row_count"]
            for measure in measure_columns:
                record[f"pbi_{measure}"] = format(left_stats["measures"][measure].normalize(), "f")
                record[f"synapse_{measure}"] = format(right_stats["measures"][measure].normalize(), "f")
            changed.append(record)

    grouping_dir = OUT_DIR / output_prefix
    write_csv(
        grouping_dir / "only_in_pbi.csv",
        only_left,
        [*key_columns, "row_count", *measure_columns],
    )
    write_csv(
        grouping_dir / "only_in_synapse.csv",
        only_right,
        [*key_columns, "row_count", *measure_columns],
    )
    write_csv(
        grouping_dir / "mismatched_groups.csv",
        changed,
        [
            *key_columns,
            "pbi_row_count",
            "synapse_row_count",
            *[f"pbi_{measure}" for measure in measure_columns],
            *[f"synapse_{measure}" for measure in measure_columns],
        ],
    )

    return {
        "key_columns": key_columns,
        "pbi_group_count": len(left_groups),
        "synapse_group_count": len(right_groups),
        "groups_only_in_pbi": len(only_left),
        "groups_only_in_synapse": len(only_right),
        "groups_with_measure_or_count_differences": len(changed),
        "only_in_pbi_preview": only_left[:10],
        "only_in_synapse_preview": only_right[:10],
        "changed_preview": changed[:10],
    }


def compare_rows_on_shared_columns(left: WorkbookData, right: WorkbookData, shared_columns: list[str]) -> dict[str, object]:
    left_counter = Counter(row_signature(row, shared_columns) for row in left.rows)
    right_counter = Counter(row_signature(row, shared_columns) for row in right.rows)

    only_left: list[dict[str, object]] = []
    only_right: list[dict[str, object]] = []
    all_signatures = sorted(set(left_counter) | set(right_counter))
    for signature in all_signatures:
        left_count = left_counter.get(signature, 0)
        right_count = right_counter.get(signature, 0)
        if left_count == right_count:
            continue

        base_record = {column: value for column, value in zip(shared_columns, signature)}
        if left_count > right_count:
            record = dict(base_record)
            record["count_difference"] = left_count - right_count
            only_left.append(record)
        else:
            record = dict(base_record)
            record["count_difference"] = right_count - left_count
            only_right.append(record)

    shared_dir = OUT_DIR / "shared_row_diff"
    write_csv(
        shared_dir / "rows_more_frequent_in_pbi.csv",
        only_left[:PREVIEW_LIMIT],
        [*shared_columns, "count_difference"],
    )
    write_csv(
        shared_dir / "rows_more_frequent_in_synapse.csv",
        only_right[:PREVIEW_LIMIT],
        [*shared_columns, "count_difference"],
    )

    return {
        "shared_column_count": len(shared_columns),
        "distinct_shared_rows_in_pbi": len(left_counter),
        "distinct_shared_rows_in_synapse": len(right_counter),
        "row_patterns_more_frequent_in_pbi": len(only_left),
        "row_patterns_more_frequent_in_synapse": len(only_right),
        "pbi_preview": only_left[:10],
        "synapse_preview": only_right[:10],
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    pbi = load_workbook_data("pbi", PBI_PATH)
    synapse = load_workbook_data("synapse", SYNAPSE_PATH)

    schema = schema_report(pbi, synapse)
    shared_columns = schema["shared_columns"]

    row_diff = compare_rows_on_shared_columns(pbi, synapse, shared_columns)
    level3 = diff_groupings(
        pbi,
        synapse,
        [COMPANY_KEY, LEVEL3_KEY],
        MEASURE_COLUMNS,
        "company_level3",
    )
    level4 = diff_groupings(
        pbi,
        synapse,
        [COMPANY_KEY, LEVEL4_KEY],
        MEASURE_COLUMNS,
        "company_level4",
    )

    summary = {
        "files": {
            "pbi": str(PBI_PATH),
            "synapse": str(SYNAPSE_PATH),
        },
        "row_counts": {
            "pbi": len(pbi.rows),
            "synapse": len(synapse.rows),
        },
        "schema": schema,
        "shared_row_comparison": row_diff,
        "aggregations": {
            "company_plus_level3": level3,
            "company_plus_level4": level4,
        },
    }

    (OUT_DIR / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
